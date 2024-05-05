import pdfplumber
import multiprocessing
import os
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime
import requests as req
from tkinter import Tk, Frame, filedialog, Label, Entry, Button, StringVar
from datetime import timedelta


def convert_to_float(valor):
    if valor == "":
        return 0.0
    return float(valor.replace(".", "").replace(",", "."))

def convert_to_int(valor):
    if valor == "":
        return 0
    return int(valor)
def credito_debito(valor, simbolo):
    if valor == "":
        return 0.0
    if valor == "0,00":
        return 0.0

    if simbolo == "C":
        return float(valor.replace(".", "").replace(",", "."))*1.0

    if simbolo == "D":
        return float(valor.replace(".", "").replace(",", "."))*-1.0
    return 0.0

   

def get_page(pdf, i):
       

    page= pdf.pages[i]
    
    
    table_negociacao_croped = page.crop((20, 250, page.width, 450))
    #table_negociacao_croped.to_image(resolution=300).draw_vlines([ 35, 43,91,105,166, 190, 305, 358, 402, 456, 543, 560])

    negocios_realizados_datas = table_negociacao_croped.extract_table({
        "vertical_strategy": "explicit",
        "horizontal_strategy": "lines",
        "explicit_vertical_lines": [35, 43,91,105,166, 190, 305, 358, 402, 456, 543, 560],
        
    })
    
    
    table_resumo_financeiro = page.crop((506, 470, page.width, 650))
    #table_resumo_financeiro.to_image(resolution=300).draw_vlines([ 515,546, 559])
    #table_resumo_financeiro.to_image(resolution=300).draw_hlines([ 480, 489, 498, 516, 525, 534, 543,  570, 579, 588, 598, 607.5, 616.5, 625, 639, 648])
    
    

    resumo_financeiro = table_resumo_financeiro.extract_table({
        "vertical_strategy": "explicit",
        "horizontal_strategy": "explicit",
        "explicit_vertical_lines": [515, 546, 559],
        "explicit_horizontal_lines": [471, 480, 489, 498, 516, 525, 534, 543,  570, 579, 588, 598, 607.5, 616.5, 625, 639, 648]
    })
    

    #page.crop((429, 60, 562, 70)).to_image(resolution=300)
    nota_folha_pregao = page.crop((429, 60, 562, 70)).extract_table()




    #page.crop((130,77, 250, 86)).to_image(resolution=300)
    corretora = page.crop((130,77, 250, 86)).extract_text()
    for papel in negocios_realizados_datas:
        valor =  {
            "nota": nota_folha_pregao[0][0],
            "folha": nota_folha_pregao[0][1],
            "data": datetime.strptime(nota_folha_pregao[0][2], "%d/%m/%Y"), 
            "q": papel[0],
            "negociacao": papel[1],
            "OP": papel[2],
            "tipo_mercado": papel[3],
            "prazo": papel[4],
            "papel": papel[5],
            "obs": papel[6],
            "quantidade": convert_to_int(papel[7]),
            "preço_ajuste": convert_to_float(papel[8]),
            "valor_operacao": convert_to_float(papel[9]),
            "d/c": papel[10],
            "valor_liquido_operacoes":credito_debito(resumo_financeiro[0][0], resumo_financeiro[0][1]),
            "taxa_liquidacao": credito_debito(resumo_financeiro[1][0], resumo_financeiro[1][1]),
            "taxa_registro": credito_debito(resumo_financeiro[2][0], resumo_financeiro[2][1]),
            "total_cblc": credito_debito(resumo_financeiro[3][0], resumo_financeiro[3][1]),
            "taxa_termo_opcoes": credito_debito(resumo_financeiro[4][0], resumo_financeiro[4][1]),
            "taxa_ana": credito_debito(resumo_financeiro[5][0], resumo_financeiro[5][1]),
            "emolumentos": credito_debito(resumo_financeiro[6][0], resumo_financeiro[6][1]),
            "taxa_operacional": credito_debito(resumo_financeiro[7][0], resumo_financeiro[7][1]),
            "execucao": credito_debito(resumo_financeiro[8][0], resumo_financeiro[8][1]),
            "taxa_custodia": credito_debito(resumo_financeiro[9][0], resumo_financeiro[9][1]),
            "impostos": credito_debito(resumo_financeiro[10][0], resumo_financeiro[10][1]),
            "irrf": credito_debito(resumo_financeiro[11][0], resumo_financeiro[11][1]),
            "outros": credito_debito(resumo_financeiro[12][0], resumo_financeiro[12][1]),
            "total_custos_despesas": credito_debito(resumo_financeiro[13][0], resumo_financeiro[13][1]),
            "liquido": credito_debito(resumo_financeiro[14][0], resumo_financeiro[14][1])
            }
        valor_resumido =  {
            "Data": valor["data"],
            "Papel": valor["papel"],
            "OP": valor["OP"],
            "QTD": valor["quantidade"],
            "Preço": valor["preço_ajuste"],
            "Valor Operação": valor["valor_operacao"],
            "Taxa de Liquidação": valor["taxa_liquidacao"],
            "Taxa de Registro": valor["taxa_registro"],
            "Taxa de Termo/Opções": valor["taxa_termo_opcoes"],
            "Taxa A.N.A": valor["taxa_ana"],
            "Emolumentos": valor["emolumentos"],
            "Taxa Operacional": valor["taxa_operacional"],
            "Execução": valor["execucao"],
            "Taxa de Custódia": valor["taxa_custodia"],
            "Impostos": valor["impostos"],
            "IRRF": valor["irrf"],
            "Outros": valor["outros"],
            "Corretora": corretora,
            "NR. Nota": valor["nota"],
            "Página": i+1}     
            
        yield valor, valor_resumido
            
def process_pdf(pdf_path):
    pdf = pdfplumber.open(pdf_path)
    pages = pdf.pages
    last_page_values = get_page(pdf, len(pages)-1).__next__()
    for i in range(len(pages)-1, -1,-1):
        for value, value_resumido in get_page(pdf, i):
            value["valor_liquido_operacoes"] = last_page_values[0]["valor_liquido_operacoes"]
            value["taxa_liquidacao"] = last_page_values[0]["taxa_liquidacao"]
            value["taxa_registro"] = last_page_values[0]["taxa_registro"]
            value["total_cblc"] = last_page_values[0]["total_cblc"]
            value["taxa_termo_opcoes"] = last_page_values[0]["taxa_termo_opcoes"]
            value["taxa_ana"] = last_page_values[0]["taxa_ana"]
            value["emolumentos"] = last_page_values[0]["emolumentos"]
            value["taxa_operacional"] = last_page_values[0]["taxa_operacional"]
            value["execucao"] = last_page_values[0]["execucao"]
            value["taxa_custodia"] = last_page_values[0]["taxa_custodia"]
            value["impostos"] = last_page_values[0]["impostos"]
            value["irrf"] = last_page_values[0]["irrf"]
            value["outros"] = last_page_values[0]["outros"]
            value["total_custos_despesas"] = last_page_values[0]["total_custos_despesas"]
            
            value_resumido["Taxa de Liquidação"] = last_page_values[1]["Taxa de Liquidação"]
            value_resumido["Taxa de Registro"] = last_page_values[1]["Taxa de Registro"]
            value_resumido["Taxa de Termo/Opções"] = last_page_values[1]["Taxa de Termo/Opções"]
            value_resumido["Taxa A.N.A"] = last_page_values[1]["Taxa A.N.A"]
            value_resumido["Emolumentos"] = last_page_values[1]["Emolumentos"]
            value_resumido["Taxa Operacional"] = last_page_values[1]["Taxa Operacional"]
            value_resumido["Execução"] = last_page_values[1]["Execução"]
            value_resumido["Taxa de Custódia"] = last_page_values[1]["Taxa de Custódia"]
            value_resumido["Impostos"] = last_page_values[1]["Impostos"]
            value_resumido["IRRF"] = last_page_values[1]["IRRF"]
            value_resumido["Outros"] = last_page_values[1]["Outros"]
            yield value, value_resumido

def get_all_files(path):
    files = os.listdir(path)
    files = [f for f in files if f.endswith(".pdf")]
    all_values = []
    all_values_resumidos = []
    for file in files:
        for value, value_resumido in process_pdf(os.path.join(path, file)):
            if value["papel"] != "":
                value["arquivo"] = file
                value_resumido["arquivo"] = file
                all_values.append(value)
                all_values_resumidos.append(value_resumido)
    return all_values_resumidos
            

            
     

def convert_to_xlsx(path, output):
    pd.DataFrame(get_all_files(path)).to_excel(output, index=False)
    wb = load_workbook(output)
    ws = wb.active
    ws.title = "Operações"
    
    fill = PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="solid")

    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            cell.fill = fill
            
    last_col_letter = get_column_letter(ws.max_column)
    for column in ws.iter_cols(min_col=5, max_col=17):
        for cell in column:
            cell.number_format = "R$ #,##0.00"
    #coluna D em formato inteiro
    for cell in ws["D"]:
        cell.number_format = "0"
    ws.auto_filter.ref = f"A1:{last_col_letter}1"
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 15
    ws.column_dimensions['Q'].width = 15
    ws.column_dimensions['R'].width = 15
    ws.column_dimensions['S'].width = 15
    ws.column_dimensions['T'].width = 15
    ws.column_dimensions['U'].width = 15
    ws.column_dimensions['V'].width = 15
    ws.column_dimensions['W'].width = 15
    ws.column_dimensions['X'].width = 15
    for cell in ws["A"]:
        cell.number_format = "dd/mm/yyyy"
        
    
    
    wb.save(output)
    print("Processamento concluído.")   
    
if __name__ == "__main__":
    
    PADX = 5
    PADY = 1    
    expiration = datetime.strptime("06/05/2024", "%d/%m/%Y")
    expiration += timedelta(hours=12)
    print(expiration)
   
    now = req.get("http://worldtimeapi.org/api/timezone/America/Sao_Paulo").json()["utc_datetime"]
    now = datetime.fromisoformat(now).replace(tzinfo=None)
    
    if expiration > now:
        """
        parser = argparse.ArgumentParser(description='Processa arquivos PDF.')
        parser.add_argument('path', type=str, help='O caminho para os arquivos PDF.')
        parser.add_argument('output', type=str, help='O caminho para o arquivo de saída.')
        args = parser.parse_args()
        """
        root = Tk()
        
        #mudar o título da janela
        root.title("Conversor de PDF para XLSX")
        frame = Frame(root)
        frame.pack(padx=25, pady=25)
        label = Label(frame, text="Caminho dos arquivos PDF:")
        label.grid(row=0, column=0, padx=PADX, pady=PADY, sticky="w")
        
        var_path = StringVar()
        entry_path = Entry(frame, width=50, state="readonly", textvariable=var_path)
        entry_path.grid(row=1, column=0, padx=PADX, pady=PADY)
        button = Button(frame, text="...", command=lambda: var_path.set(filedialog.askdirectory()))
        button.grid(row=1, column=1, padx=PADX, pady=PADY)
        
        Label(frame, text="Caminho do arquivo de saída:").grid(row=2, column=0, padx=PADX, pady=PADY, sticky="w")
        var_output = StringVar()
        entry_output = Entry(frame, width=50, state="readonly" ,textvariable=var_output)
        entry_output.grid(row=3, column=0, padx=PADX, pady=PADY)
        
        button_output = Button(frame, text="...", command=lambda: var_output.set(filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])))
        button_output.grid(row=3, column=1, padx=PADX, pady=PADY)
        
        
        button_convert = Button(frame, text="Submeter", command=lambda: convert_to_xlsx(var_path.get(), var_output.get()))
        button_convert.grid(row=4, column=0, padx=PADX, pady=(PADY*3, 0))
        
        
        root.mainloop()
        
        
        

    else:
        print("Licença expirada")
    